#!/usr/bin/env python3
"""
SEL (Slowly Expanding Lesion) lateralization workflow.

Pipeline
--------
1. Load a subject's T1 MRI, SEL mask, and (optionally) baseline lesion mask (.nii/.nii.gz).
2. Register the T1 to MNI152 space (affine), carrying the masks along with the
   SAME transform using nearest-neighbour interpolation (labels are preserved).
3. In MNI space, the mid-sagittal plane is x = 0 (the anterior-commissure origin),
   so "cut the brain in half vertically in the transverse/axial view" == split on
   MNI x-coordinate. Each connected SEL component is counted as ONE lesion and
   assigned to Left / Right / Middle based on its centroid (Middle = straddles
   the midline within a tolerance).
4. Write a per-subject + summary Excel file matching the provided header, and
   (bonus) save axial overlay PNGs and a bar chart.

MNI convention used: voxel/world x increases to the SUBJECT's RIGHT only if the
image is stored RAS+. nibabel reports orientation from the affine, so we map
each lesion centroid through the affine to world (MNI) coordinates and use the
sign of world-x. In MNI152: world-x > 0  -> Right hemisphere,
                          world-x < 0  -> Left hemisphere,
                          |world-x| <= tol -> Middle (crosses midline).

Usage
-----
Most users should run the FSL wrapper (run_sel_fsl.sh) instead of calling this
directly. To call directly:

Single subject (built-in Python registration):
    python sel_lateralization.py \
        --case-id SUBJ001 \
        --t1   t1.nii.gz \
        --sel  sel_mask.nii.gz \
        --baseline baseline_mask.nii.gz \
        --out  ./results

Batch (CSV with columns: CaseId,T1,SEL_mask,Lesion_baseline):
    python sel_lateralization.py --manifest cases.csv --out ./results

Inputs already registered to MNI by FSL FLIRT (count without resampling):
    add --no-resample
"""

import argparse
import os
import sys
import numpy as np
import nibabel as nib
from scipy import ndimage


# --------------------------------------------------------------------------- #
# Registration
# --------------------------------------------------------------------------- #
def register_to_mni(t1_img, mask_imgs, skip=False):
    """
    Affine-register T1 to the MNI152 template and resample masks with the same
    transform (nearest-neighbour). Returns (t1_mni, [mask_mni, ...]).

    If skip=True, assumes inputs are already in MNI space and only resamples
    onto the MNI152 grid so coordinates are interpretable.
    """
    from nilearn.datasets import load_mni152_template
    from nilearn.image import resample_to_img, resample_img

    template = load_mni152_template(resolution=1)

    if skip:
        # Just put everything on the template grid; affine already MNI.
        t1_mni = resample_to_img(t1_img, template, interpolation="continuous",
                                 force_resample=True, copy_header=True)
        masks_mni = [
            resample_to_img(m, template, interpolation="nearest",
                            force_resample=True, copy_header=True)
            for m in mask_imgs
        ]
        return t1_mni, masks_mni

    # --- Affine registration of T1 -> template via nilearn/scipy ---
    # nilearn has no built-in optimiser; we use a robust intensity-based
    # affine from the `nilearn`-bundled approach: resample T1 to template grid
    # (which applies the header affine), then refine with a mutual-information
    # affine using scipy's optimizer.
    t1_on_grid = resample_to_img(t1_img, template, interpolation="continuous",
                                 force_resample=True, copy_header=True)

    moving = np.nan_to_num(t1_on_grid.get_fdata().astype(np.float32))
    fixed = np.nan_to_num(template.get_fdata().astype(np.float32))

    affine_params = _affine_register(fixed, moving)
    t1_mni_data = _apply_affine(moving, affine_params, order=1)
    t1_mni = nib.Nifti1Image(t1_mni_data, template.affine, template.header)

    masks_mni = []
    for m in mask_imgs:
        m_on_grid = resample_to_img(m, template, interpolation="nearest",
                                    force_resample=True, copy_header=True)
        mdata = m_on_grid.get_fdata()
        warped = _apply_affine(mdata.astype(np.float32), affine_params, order=0)
        warped = (warped > 0.5).astype(np.uint8)
        masks_mni.append(nib.Nifti1Image(warped, template.affine, template.header))

    return t1_mni, masks_mni


def _affine_register(fixed, moving, n_iter=200):
    """
    Lightweight 12-param affine registration by Powell optimisation on negative
    normalised cross-correlation. Operates on downsampled volumes for speed.
    Returns a 4x4 matrix mapping fixed-grid coords -> moving-grid coords.
    """
    from scipy.optimize import minimize

    ds = 4  # downsample factor for the optimiser
    f = fixed[::ds, ::ds, ::ds]
    f = (f - f.mean()) / (f.std() + 1e-8)

    center = (np.array(moving.shape) - 1) / 2.0

    def cost(p):
        mat = _params_to_matrix(p, center, ds)
        warped = ndimage.affine_transform(
            moving, mat[:3, :3], offset=mat[:3, 3],
            output_shape=f.shape, order=1, mode="constant", cval=0.0)
        w = (warped - warped.mean()) / (warped.std() + 1e-8)
        return -float(np.mean(f * w))

    x0 = np.zeros(12)
    x0[6:9] = 1.0  # initial scales = 1
    res = minimize(cost, x0, method="Powell",
                   options={"maxiter": n_iter, "xtol": 1e-3, "ftol": 1e-3})
    # Rebuild full-resolution matrix (ds=1)
    return _params_to_matrix(res.x, center, 1)


def _params_to_matrix(p, center, ds):
    """12 params (3 trans, 3 rot, 3 scale, 3 shear) -> 4x4 sampling matrix."""
    tx, ty, tz, rx, ry, rz, sx, sy, sz, hxy, hxz, hyz = p
    sx, sy, sz = abs(sx) + 1e-3, abs(sy) + 1e-3, abs(sz) + 1e-3

    Rx = np.array([[1, 0, 0], [0, np.cos(rx), -np.sin(rx)], [0, np.sin(rx), np.cos(rx)]])
    Ry = np.array([[np.cos(ry), 0, np.sin(ry)], [0, 1, 0], [-np.sin(ry), 0, np.cos(ry)]])
    Rz = np.array([[np.cos(rz), -np.sin(rz), 0], [np.sin(rz), np.cos(rz), 0], [0, 0, 1]])
    R = Rz @ Ry @ Rx
    S = np.diag([sx, sy, sz])
    H = np.array([[1, hxy, hxz], [0, 1, hyz], [0, 0, 1]])
    A = R @ S @ H

    mat = np.eye(4)
    mat[:3, :3] = A
    # rotate about volume centre, then translate
    mat[:3, 3] = center - A @ center + np.array([tx, ty, tz])
    # account for downsampling of the fixed grid
    scale = np.eye(4)
    scale[0, 0] = scale[1, 1] = scale[2, 2] = ds
    return mat @ scale


def _apply_affine(vol, mat, order=1):
    return ndimage.affine_transform(
        vol, mat[:3, :3], offset=mat[:3, 3],
        output_shape=vol.shape, order=order, mode="constant", cval=0.0)


# --------------------------------------------------------------------------- #
# Lesion counting / lateralization
# --------------------------------------------------------------------------- #
def count_lesions_by_side(sel_img, midline_tol_mm=2.0, min_voxels=1):
    """
    Label connected components in the SEL mask and assign each to a hemisphere
    using its centroid's world-x (MNI) coordinate.

    Returns dict with counts and a per-lesion table.
    """
    data = sel_img.get_fdata()
    affine = sel_img.affine
    mask = data > 0.5

    # 26-connectivity so a lesion isn't split by diagonal voxels
    structure = np.ones((3, 3, 3), dtype=int)
    labels, n = ndimage.label(mask, structure=structure)

    lesions = []
    left = right = middle = 0
    for lbl in range(1, n + 1):
        coords = np.argwhere(labels == lbl)
        if coords.shape[0] < min_voxels:
            continue
        centroid_vox = coords.mean(axis=0)
        world = affine @ np.array([*centroid_vox, 1.0])
        wx = world[0]
        if abs(wx) <= midline_tol_mm:
            side = "Middle"; middle += 1
        elif wx > 0:
            side = "Right"; right += 1
        else:
            side = "Left"; left += 1
        lesions.append({
            "lesion_id": lbl,
            "voxels": int(coords.shape[0]),
            "centroid_x_mm": round(float(wx), 2),
            "centroid_y_mm": round(float(world[1]), 2),
            "centroid_z_mm": round(float(world[2]), 2),
            "side": side,
        })

    return {
        "n_total": len(lesions),
        "n_left": left,
        "n_right": right,
        "n_middle": middle,
        "lesions": lesions,
    }


# --------------------------------------------------------------------------- #
# Visualization (bonus)
# --------------------------------------------------------------------------- #
def save_overlay(t1_img, sel_img, out_png, case_id):
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    # Images are already RAS-canonical at this point (axis 0=L-R, 1=P-A, 2=I-S).
    t1 = t1_img.get_fdata()
    sel = sel_img.get_fdata() > 0.5
    aff = t1_img.affine

    # MNI x=0 midline in voxel space
    inv = np.linalg.inv(aff)
    x0_vox = int(round((inv @ np.array([0, 0, 0, 1]))[0]))

    # Pick axial slices (axis 2) where lesions exist
    sel_z = np.where(sel.any(axis=(0, 1)))[0]
    if sel_z.size:
        slices = np.linspace(sel_z.min(), sel_z.max(), min(6, sel_z.size)).astype(int)
    else:
        slices = np.linspace(t1.shape[2] * 0.3, t1.shape[2] * 0.7, 6).astype(int)

    # In RAS canonical: axial slice = t1[:, :, z], displayed with
    # rot90 to put anterior up; midline is a vertical line at x0_vox.
    fig, axes = plt.subplots(2, 3, figsize=(12, 8))
    for ax, z in zip(axes.ravel(), slices):
        ax.imshow(np.rot90(t1[:, :, z]), cmap="gray")
        overlay = np.ma.masked_where(~sel[:, :, z], sel[:, :, z])
        ax.imshow(np.rot90(overlay), cmap="autumn", alpha=0.8)
        # rot90 flips axis-0 to become columns, so midline x-voxel maps to
        # a vertical line at (shape[0]-1 - x0_vox) after rotation
        midline_col = t1.shape[0] - 1 - x0_vox
        ax.axvline(midline_col, color="cyan", lw=1, ls="--")
        ax.set_title(f"axial z={z}", fontsize=9)
        ax.axis("off")
    fig.suptitle(f"{case_id}: SEL overlay (cyan = MNI midline)", fontsize=13)
    fig.tight_layout()
    fig.savefig(out_png, dpi=110, bbox_inches="tight")
    plt.close(fig)


def save_barchart(rows, out_png):
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    cases = [r["CaseId"] for r in rows]
    left = [r["SEL_left"] for r in rows]
    right = [r["SEL_right"] for r in rows]
    mid = [r["SEL_middle"] for r in rows]
    x = np.arange(len(cases)); w = 0.25

    fig, ax = plt.subplots(figsize=(max(6, len(cases) * 1.2), 5))
    ax.bar(x - w, left, w, label="Left", color="#4C72B0")
    ax.bar(x, right, w, label="Right", color="#C44E52")
    ax.bar(x + w, mid, w, label="Middle", color="#999999")
    ax.set_xticks(x); ax.set_xticklabels(cases, rotation=45, ha="right")
    ax.set_ylabel("SEL count"); ax.set_title("SEL lateralization by case")
    ax.legend()
    fig.tight_layout()
    fig.savefig(out_png, dpi=120, bbox_inches="tight")
    plt.close(fig)


# --------------------------------------------------------------------------- #
# Excel output
# --------------------------------------------------------------------------- #
def write_excel(rows, out_xlsx):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "SEL_lateralization"

    headers = ["CaseId", "T1", "SEL_mask", "SEL_number", "Lesion_baseline",
               "SEL_left", "SEL_right", "SEL_middle"]
    ws.append(headers)
    hfill = PatternFill("solid", start_color="2F5496")
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial")
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center")

    for i, r in enumerate(rows):
        excel_row = i + 2
        ws.append([
            r["CaseId"], r["T1"], r["SEL_mask"], r["SEL_number"],
            r["Lesion_baseline"], r["SEL_left"], r["SEL_right"], r["SEL_middle"],
        ])
        # SEL_number as a checking formula = left+right+middle
        ws.cell(row=excel_row, column=4).value = f"=F{excel_row}+G{excel_row}+H{excel_row}"

    # totals
    tot_row = len(rows) + 2
    ws.cell(row=tot_row, column=1, value="TOTAL").font = Font(bold=True, name="Arial")
    for col in ("D", "E", "F", "G", "H"):
        ws.cell(row=tot_row, column={"D":4,"E":5,"F":6,"G":7,"H":8}[col],
                value=f"=SUM({col}2:{col}{tot_row-1})").font = Font(bold=True, name="Arial")

    widths = {"A": 16, "B": 28, "C": 28, "D": 12, "E": 16, "F": 10, "G": 10, "H": 11}
    for col, wdt in widths.items():
        ws.column_dimensions[col].width = wdt

    wb.save(out_xlsx)


# --------------------------------------------------------------------------- #
# Driver
# --------------------------------------------------------------------------- #
def process_case(case_id, t1_path, sel_path, baseline_path, out_dir,
                 skip_registration=False, make_overlay=True, no_resample=False):
    t1_img = nib.load(t1_path)
    sel_img = nib.load(sel_path)
    mask_imgs = [sel_img]
    has_baseline = baseline_path and os.path.exists(baseline_path)
    if has_baseline:
        mask_imgs.append(nib.load(baseline_path))

    if no_resample:
        # Inputs are ALREADY registered to MNI by an external tool (e.g. FSL
        # FLIRT). Reorient to RAS canonical so axis-2 is always axial and
        # world-x sign is consistent for L/R assignment.
        t1_mni = nib.as_closest_canonical(t1_img)
        sel_mni = nib.as_closest_canonical(sel_img)
        baseline_mni = nib.as_closest_canonical(mask_imgs[1]) if has_baseline else None
    else:
        t1_mni, masks_mni = register_to_mni(t1_img, mask_imgs, skip=skip_registration)
        t1_mni = nib.as_closest_canonical(t1_mni)
        sel_mni = nib.as_closest_canonical(masks_mni[0])
        baseline_mni = nib.as_closest_canonical(masks_mni[1]) if has_baseline else None

        # save MNI-space niftis (only when we produced them)
        nib.save(t1_mni, os.path.join(out_dir, f"{case_id}_T1_MNI.nii.gz"))
        nib.save(sel_mni, os.path.join(out_dir, f"{case_id}_SEL_MNI.nii.gz"))
        if baseline_mni is not None:
            nib.save(baseline_mni, os.path.join(out_dir, f"{case_id}_baseline_MNI.nii.gz"))

    counts = count_lesions_by_side(sel_mni)

    baseline_count = ""
    if baseline_mni is not None:
        bc = count_lesions_by_side(baseline_mni)
        baseline_count = bc["n_total"]

    if make_overlay:
        try:
            save_overlay(t1_mni, sel_mni,
                         os.path.join(out_dir, f"{case_id}_SEL_overlay.png"), case_id)
        except Exception as e:
            print(f"[warn] overlay failed for {case_id}: {e}", file=sys.stderr)

    return {
        "CaseId": case_id,
        "T1": os.path.basename(t1_path),
        "SEL_mask": os.path.basename(sel_path),
        "SEL_number": counts["n_total"],
        "Lesion_baseline": baseline_count,
        "SEL_left": counts["n_left"],
        "SEL_right": counts["n_right"],
        "SEL_middle": counts["n_middle"],
        "_lesions": counts["lesions"],
    }


def main():
    ap = argparse.ArgumentParser(description="SEL lateralization workflow")
    ap.add_argument("--case-id")
    ap.add_argument("--t1")
    ap.add_argument("--sel")
    ap.add_argument("--baseline")
    ap.add_argument("--manifest", help="CSV: CaseId,T1,SEL_mask,Lesion_baseline")
    ap.add_argument("--out", default="./results")
    ap.add_argument("--skip-registration", action="store_true",
                    help="Inputs already in MNI space (resample onto template grid)")
    ap.add_argument("--no-resample", action="store_true",
                    help="Inputs already registered to MNI by an external tool "
                         "(e.g. FSL FLIRT); count on the native grid without "
                         "resampling. Use this for the FSL workflow.")
    ap.add_argument("--no-overlay", action="store_true")
    args = ap.parse_args()

    os.makedirs(args.out, exist_ok=True)
    rows = []

    if args.manifest:
        import csv
        with open(args.manifest) as f:
            for r in csv.DictReader(f):
                rows.append(process_case(
                    r["CaseId"], r["T1"], r["SEL_mask"],
                    r.get("Lesion_baseline") or None, args.out,
                    skip_registration=args.skip_registration,
                    make_overlay=not args.no_overlay,
                    no_resample=args.no_resample))
    else:
        if not (args.case_id and args.t1 and args.sel):
            ap.error("Provide --manifest OR (--case-id --t1 --sel)")
        rows.append(process_case(
            args.case_id, args.t1, args.sel, args.baseline, args.out,
            skip_registration=args.skip_registration,
            make_overlay=not args.no_overlay,
            no_resample=args.no_resample))

    # per-lesion detail csv
    import csv as _csv
    detail_path = os.path.join(args.out, "SEL_lesion_detail.csv")
    with open(detail_path, "w", newline="") as f:
        w = _csv.writer(f)
        w.writerow(["CaseId", "lesion_id", "voxels", "centroid_x_mm",
                    "centroid_y_mm", "centroid_z_mm", "side"])
        for row in rows:
            for L in row["_lesions"]:
                w.writerow([row["CaseId"], L["lesion_id"], L["voxels"],
                            L["centroid_x_mm"], L["centroid_y_mm"],
                            L["centroid_z_mm"], L["side"]])

    for r in rows:
        r.pop("_lesions", None)

    out_xlsx = os.path.join(args.out, "SEL_lateralization_results.xlsx")
    write_excel(rows, out_xlsx)
    if not args.no_overlay:
        try:
            save_barchart(rows, os.path.join(args.out, "SEL_lateralization_chart.png"))
        except Exception as e:
            print(f"[warn] chart failed: {e}", file=sys.stderr)

    print("Done.")
    for r in rows:
        print(f"  {r['CaseId']}: total={r['SEL_number']} "
              f"L={r['SEL_left']} R={r['SEL_right']} M={r['SEL_middle']}")
    print(f"Excel: {out_xlsx}")


if __name__ == "__main__":
    main()
